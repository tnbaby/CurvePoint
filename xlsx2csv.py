import argparse
from openpyxl import load_workbook
import csv

parser = argparse.ArgumentParser()

parser.add_argument('-input', default = '', help = 'input file')
parser.add_argument('-output', default = '', help = 'output file')

args = parser.parse_args()

# read excel file from workbook
wb = load_workbook(filename=args.input)

sheets = wb.get_sheet_names()
sheet_first = sheets[0]
ws = wb.get_sheet_by_name(sheet_first)

rows = ws.rows
columns = ws.columns

data = []
for row in rows:
	line = [col.value for col in row]
	data += [line]

print ws.cell('A1').value
print ws.cell(row=1, column=1).value

# write data to csv file
csv_writer = csv.writer(open(args.output, 'w'))
csv_writer.writerow(data)
